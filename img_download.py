import openpyxl
import requests
from PIL import Image
from io import BytesIO

def download_and_resize_image(url, size=(800, 900)):
    response = requests.get(url, timeout=10)  # Add a timeout
    response.raise_for_status()  # Raise an exception for bad responses
    img = Image.open(BytesIO(response.content))
    img = img.resize(size, Image.LANCZOS)
    img_byte_arr = BytesIO()
    img.save(img_byte_arr, format='PNG')
    return img_byte_arr.getvalue()

def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):  # Assuming the first row is header
        url = sheet.cell(row=row, column=1).value
        if url:
            try:
                img_data = download_and_resize_image(url)
                img = openpyxl.drawing.image.Image(BytesIO(img_data))
                img.width, img.height = 800, 900  # Set image size in Excel
                cell = sheet.cell(row=row, column=12)  # Column L
                cell.value = None  # Clear any existing value
                img.anchor = cell.coordinate
                sheet.add_image(img)
                print(f"Successfully processed row {row}")
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                # Leave column L blank for invalid links
                sheet.cell(row=row, column=12).value = None
                # Continue with the next row
                continue

    wb.save(file_path)

# Usage
excel_file = "f2.xlsx"
process_excel(excel_file)