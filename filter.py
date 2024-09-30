import openpyxl

def extract_dimensions(file_path):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Column L is the 12th column (A=1, B=2, ..., L=12)
    dimension_column = 15

    # Add the header for the Dimension column
    sheet.cell(row=1, column=dimension_column, value="Pattern")

    # Iterate through all rows
    for row in range(2, sheet.max_row + 1):  # Start from 2 to skip header row
        dimension_found = False
        
        # Check each cell in the row
        for cell in sheet[row]:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("Pattern"):
                # Copy the value to column L
                sheet.cell(row=row, column=dimension_column, value=cell.value)
                dimension_found = True
                break
        
        # If no dimension found, the cell in column L remains empty

    # Save the modified workbook
    workbook.save(file_path)
    print(f"Pattern extracted and saved to column O")

# Usage
file_path = "vidhan_1.xlsx"
extract_dimensions(file_path)